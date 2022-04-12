from openpyxl import Workbook,load_workbook
from collections import defaultdict
import pandas as pd
import time
import math
import re
import copy



# Convert the location and slot cells of excel i.e x_coord,y_coord and slot_start,slot_end
def preprocess_loc_slot_xl(x,y):
    return [int(x),int(y)]

# Convert the skills cell to list of skills and trim the extra spaces from each skill(assumes skill data in lowercase)
def preprocess_skills_xl(skill_str,requiredSkills = None):
    skills = skill_str.split(',')
    fin_skill = set()
    for i in range(len(skills)):
        s = re.sub(' +',' ',skills[i].strip())
        if requiredSkills and s in requiredSkills:
            fin_skill.add(s)
        elif requiredSkills is None:
            fin_skill.add(s)
    return fin_skill


class MainTask(object):
    def __init__(self,t_name,t_group):
        self.name = t_name
        self.group = t_group

class Task(object):
    def __init__(self,t_name,t_id,t_skills,t_loc,budget,preference):
        self.name = t_name
        self.id = t_id
        self.skills = t_skills
        self.location = t_loc
        self.skillsLeft = len(self.skills)
        self.finished = False
        self.budget = budget
        self.tempPreference = preference

    def fillPreference(self,NameVolunteersMap):
        self.preference = {}
        rank = 1
        for a_name in self.tempPreference:
            self.preference[NameVolunteersMap[a_name]] = rank
            rank+=1

class Volunteer(object):
    def __init__(self,a_name,a_id,a_skills,a_loc,a_slot,remuneration,preference):
        self.name = a_name
        self.id = a_id
        self.skills = a_skills
        self.location = a_loc
        self.slot = a_slot
        self.remuneration = remuneration
        self.numSkills = len(self.skills)
        self.assigned = False
        self.utilisedSkills = set()
        self.costIncured = None
        self.assignedTask = None
        self.tempPreference = preference

    def fillPreference(self,NameTaskMap):
        self.preference = {}
        rank = 1
        for t_name in self.tempPreference:
            self.preference[NameTaskMap[t_name]] = rank
            rank+=1

class VolunteerTaskMapping(object):
    def __init__(self,taskObject,volunteersInfo):
        self.task = taskObject
        self.groupInfo = volunteersInfo
        self.parseGroupInfo()
        self.groupSize = len(self.groupInfo)
        # if self.task.finished:
        #     self.parseSlots()
        #     self.getCommonSlots()
        #     if self.common_slots == {}:
        #         self.task.finished = False
        #     else:
        #         self.proposed_Slot = self.common_slots[max(self.common_slots)]

    def parseGroupInfo(self):
        self.volunteers = set()
        for volunteer,skills,costIncured in self.groupInfo:
            volunteer.assigned = True
            volunteer.utilisedSkills = skills
            volunteer.costIncured = costIncured
            volunteer.assignedTask = self.task
            self.volunteers.add(volunteer)

    # Function to preprocess the original time slots HHMM -> MMMM
    def preprocessLocation(self,slot):
        s_hour = slot[0]//100
        s_minutes = slot[0]%100
        start = s_hour*60+s_minutes

        # Convert end time to minutes : HHMM -> 60*HH+MM
        e_hour = slot[1]//100
        e_minutes = slot[1]%100
        end = e_hour*60+e_minutes

        return [start,end]

    def parseSlots(self):
        self.slots = []
        for volunteer in self.volunteers:
            self.slots.append(self.preprocessLocation(volunteer.slot))


    def computeFinalSlot(self,time,thresh):
        # Stores the final slot with maximum time
        start = -1
        end = -1

        # Pointers to check the current maximum slot
        i = -1
        j = -1

        # To compare the next valid slot with the previous maximum
        curMaxTime = 0 

        # Loop through the time array in order to get the valid slot and update the final slot start and end if end-start+1 > curMaxTime
        while i<len(time):
            if time[i]>=thresh:
                j = i+1
                while j<len(time) and time[j]>=thresh:
                    j+=1
                if j-i>curMaxTime:
                    start = i
                    end = j-1
                    curMaxTime = j-i
                if j<len(time):
                    i = j
                else:
                    break
            else:
                i+=1

        # Re conver the start and end to its original form i.e MMMM -> HHMM
        if end != -1 and start != -1:
            if end-start<60:
                if start>60:
                    start-=60
                if end<1379:
                    end+=60
            s_hour = start//60
            s_minutes = start%60
            start = s_hour*100+s_minutes

            e_hour = end//60
            e_minutes = end%60
            end = e_hour*100+e_minutes

            return [start,end]
        return []


    # Function to compute the threshold for freelancers
    def getThreshVolunteers(self):
        n = len(self.slots)

        # Initial threshold percentage = 100%
        thresh = 1

        # Set to store the number of freelancers from range [50% * n to 100% * n]
        n_thresh = set()

        # Compute the number of freelancers as per threshold eg 50% 10 = 5
        while thresh>=0.5:
            n_thresh.add(int(thresh*n))
            thresh-=0.1
        return n_thresh


    # Function to compute the common time slot between all freelancers
    def getCommonSlots(self):

        # Total minutes in a day can be from 0000 : 12AM to 1439 : 11:59PM
        time_slot_minutes = [0 for i in range(24*60+1)]

        # For each slot just mark the start marker i.e +=1 and the end marker i.e -=1 : T.C = O(len(slots))
        for slot in self.slots:
            # Mark the start time for given slot
            time_slot_minutes[slot[0]]+=1

            if slot[1]+1<len(time_slot_minutes):
                # Mark the end time for given slot
                time_slot_minutes[slot[1]+1]-=1


        # Prefix sum of the time_slot_minutes array in order to compute the count of slots in which that minute appeared.
        # Example if minute 15 -> 0015AM was a part of 10 queries then time_slot_minutes[15] = 10
        for i in range(1,len(time_slot_minutes)):
            time_slot_minutes[i]+=time_slot_minutes[i-1]

        # To store all the common slots found for the given number of threshold free lancers as key
        self.common_slots = {}

        # Number of threshold freelancers to look for [60% of total freelancers to 100% of total freelancers]
        n_thresh = self.getThreshVolunteers()

        # Check if the common slot is available for the calculated thresholds
        for threshold in n_thresh:
            commonslot = self.computeFinalSlot(time_slot_minutes,threshold)

            # If available print and return
            if commonslot != []:
                # print(f"The common slot which is common among {threshold} freelancers is : {commonslot}.")
                self.common_slots[threshold] = commonslot

class ScoreVTM(object):
    def __init__(self,VTM,taskCount,volunteers,mainTaskObjects):
        self.VTM_Objects = VTM
        self.totalTasks = taskCount
        self.Volunteers = volunteers
        self.mainTaskObjects = mainTaskObjects

    def computeSuccessRatio(self):
        count = 0
        for vtm in self.VTM_Objects:
            if vtm.task.finished:
                count+=1
        self.successRatio = count/self.totalTasks

    def computeTotalSatisfactoryRate(self):
        TSR = 0
        assignedApplicants = 0
        for vtm in self.VTM_Objects:
            if vtm.task.finished:
                assignedApplicants+=vtm.groupSize
                minute_slots = vtm.slots
                minute_proposed_Slot = vtm.preprocessLocation(vtm.proposed_Slot)


                res = 0
                start2,end2 = minute_proposed_Slot[0],minute_proposed_Slot[1]
                # print(f"Start2 :{start2}\tEnd2 : {end2}")
                for start1,end1 in minute_slots:
                    # print(f"Start1 :{start1}\tEnd1 : {end1}")
                    temp = 0
                    if start1>end2 or start2>end1:
                        continue
                    else:
                        temp = (min(end2,end1)-max(start2,start1)+1)/(end2-start2+1)
                        res+=temp
                    # print(f"calculated satisfactoryRate : {temp}")
                TSR+=res 

        TSR/=assignedApplicants
        self.totalSatisfactoryRate_TSR = TSR
        return self.totalSatisfactoryRate_TSR

    def computeNetUtilityScore(self):
        self.utilityScoreDict = dict(zip([volunteer.name for volunteer in self.Volunteers],[0 for i in range(len(self.Volunteers))]))
        self.NetUtilityScore = 0 
        for volunteer in self.Volunteers:
            if volunteer.assigned:
                skillsMatched = len(volunteer.utilisedSkills)
                costIncured = volunteer.costIncured
                self.utilityScoreDict[volunteer.name]+=(skillsMatched/costIncured)
                self.NetUtilityScore+=self.utilityScoreDict[volunteer.name]
        return self.utilityScoreDict,self.NetUtilityScore 

    def computeOverallCompletionRate_OCR(self):
        self.totalCompletionRatio = 0
        for mainTask in self.mainTaskObjects:
            score  = 0 
            for subtask in mainTask.group:
                if subtask.finished:
                    score+=1
            self.totalCompletionRatio+= (score/len(mainTask.group))
        self.OCR = self.totalCompletionRatio/len(self.mainTaskObjects)
        return self.OCR

    def getScores(self):
        self.computeSuccessRatio()
        # self.computeTotalSatisfactoryRate()
        self.computeNetUtilityScore()
        self.computeOverallCompletionRate_OCR()
        self.scores = {}
        self.scores["Success_Ratio"] = self.successRatio
        # self.scores["Total_Satisfactory_Rate"] = self.totalSatisfactoryRate_TSR
        self.scores["Net_Utility_Score"] = self.NetUtilityScore
        self.scores["Overall_Completion_Rate"] = self.OCR
        return self.scores


# Convert thr MainTaskInfo.xlsx data row by row to MainTask dictionary 
def getMainTasksObjects(filename,NameTaskMap):
    wb = load_workbook(filename)
    ws = wb.active

    T = set()
    for row in ws.iter_rows(min_row = 2,max_row = 7,min_col = 1,max_col = 2,values_only = True):
        name = row[0]
        subTaskObjects = []
        subtasks = row[1].split(",")
        for subtask in subtasks:
            subTaskObjects.append(NameTaskMap[subtask])
        obj = MainTask(name,subTaskObjects)
        T.add(obj)
    return T

# Convert thr Tasks.xlsx data row by row to Task Objects
def getTaskObjects(filename):
    wb = load_workbook(filename)
    ws = wb.active

    T = set()
    Name_Obj = {}
    t_id = 0
    for row in ws.iter_rows(min_row = 2,max_row = 7,min_col = 1,max_col = 6,values_only = True):
        skills = preprocess_skills_xl(row[1])
        location = preprocess_loc_slot_xl(row[2],row[3])
        budget = int(row[4])
        preference = row[5].split(">")
        obj = Task(row[0],t_id,skills,location,budget,preference)
        T.add(obj)
        Name_Obj[obj.name] = obj
        t_id+=1

    return T,Name_Obj

# Convert volunteer data row by row to volunteer Objects
def getVolunteerObjects(filename,requiredSkills):
    wb = load_workbook(filename)
    ws = wb.active

    V = []
    Name_Obj = {}
    a_id = 0
    for row in ws.iter_rows(min_row = 2,max_row = 14,min_col = 1,max_col = 8,values_only = True):
        skills = preprocess_skills_xl(row[1],requiredSkills)
        location = preprocess_loc_slot_xl(row[2],row[3])
        slot = preprocess_loc_slot_xl(row[4],row[5])
        remuneration = int(row[6])
        preference = row[7].split(">")
        obj = Volunteer(row[0],a_id,skills,location,slot,remuneration,preference)
        V.append(obj)
        Name_Obj[obj.name] = obj
        a_id+=1

    V.sort(key = lambda i : i.numSkills,reverse = True)

    return V,Name_Obj

# Fill the preference for task and volunteer objects
def fillPreference(Tasks,NameTaskMap,Volunteers,NameVolunteersMap):
    for task in Tasks:
        task.fillPreference(NameVolunteersMap)

    for volunteer in Volunteers:
        volunteer.fillPreference(NameTaskMap)

def getVTMObjects(VTM):
    VTM_Objects = set()
    for task,volunteersInfo in VTM.items():
        VTM_Objects.add(VolunteerTaskMapping(task,volunteersInfo))

    return VTM_Objects

# Returns dictionary of the skills required in all tasks with id attached to each skill
def getRequiredSkills(Tasks):
    S = {}
    s_id = 0
    for task in Tasks:
        for skill in task.skills:
            if S.get(skill,-1) == -1:
                S[skill] = s_id
                s_id+=1
    return S



'''
Phase 1 functions : Volunteer task mapping algorithm
'''
# Generate G_ST -- O(T*Skills)
def generateSkillTaskMapper(taskObjects,requiredSkills):

    # Number of skills left to be sufficed
    assignmentsLeft = [0]

    # Iterate on task T to set the counts on the desired indices of G_ST matrix
    G_ST = [[0 for j in range(len(taskObjects))]for i in range(len(requiredSkills))]
    for task in taskObjects:
        j = task.id 
        for skill in task.skills:
            i = requiredSkills[skill]
            G_ST[i][j] = 1
            assignmentsLeft[0]+=1
    return G_ST,assignmentsLeft

# Finds the tasks which match the skillset of a particular applicant to the maximum and returns {task : matched skills} dict, : O(T*Skills)
def getTopMatchedTasks(volunteerSkills,incompleteTaskObjects,requiredSkills,G_ST):

    recommend = {}

    matchedSkills = dict(zip(list(incompleteTaskObjects),[set() for i in range(len(incompleteTaskObjects))]))
    maximumMatchedSkills = 0

    for task in incompleteTaskObjects:
        j = task.id
        for skill in volunteerSkills:
            i = requiredSkills[skill]
            if G_ST[i][j]>0:
                matchedSkills[task].add(skill)
                maximumMatchedSkills = max(len(matchedSkills[task]),maximumMatchedSkills)

    # If none of the skill matched with any task return empty dict
    if maximumMatchedSkills == 0:
        return {}

    # If skills matched then find the tasks for whom maximum skills matched and populate the set of skills which actually matched
    for key,val in matchedSkills.items():
        if len(val) == maximumMatchedSkills:
            recommend[key] = val

    return recommend

# Compute the nearest task among all the given tasks for the given applicant location : O(T)
def getCostEfficientTask(volunteerLocation,tasks,costPerKM):

    minCost = float('inf')
    minCostTask = None

    for task in tasks:
        source = volunteerLocation
        destination = task.location

        distance = math.sqrt((destination[0]-source[0])**2+(destination[1]-source[1])**2)
        costIncured = distance*costPerKM

        if costIncured<minCost:
            minCost = costIncured
            minCostTask = task

    return minCostTask,minCost


def getMostWillingTask(volunteer,tasks):
    maxWillingnesss = -1
    maxWillingnessTask = None 

    for task,skills in tasks.items():
        rank = volunteer.preference[task]
        willingness = len(skills)/rank

        if willingness>maxWillingnesss:
            maxWillingnesss = willingness
            maxWillingnessTask = task

    return maxWillingnessTask


# Updates the G_ST as per the skills fulfilled by the given applicant : O(Skills)
def updateGST(task,skills,requiredSkills,G_ST):
    for skill in skills:
        s_idx = requiredSkills[skill]
        G_ST[s_idx][task.id]-=1


'''
Function to generate VTM (Task -> Volunteer mapping)
'''
def generateVolunteerTaskMap(G_ST,taskObjects,volunteerObjects,requiredSkills,costPerKM,assignmentsLeft):

    incompleteTaskObjects = set(taskObjects)

    completedTaskObjects = set()

    # The final map containing each task as key and the list of names of candidates assigned to that task as value
    volunteerTaskMap = defaultdict(list)

    # Iterate over the volunteers to assign it to a particular task : O(A*(T*Skills + Skills + T)) = O(T*A*Skills)
    for volunteer in volunteerObjects:

        # Finds the tasks which match the skillset of a particular applicant to the maximum and returns {task : matched skills} dict, : O(T*Skills)
        suggestedTasks = getTopMatchedTasks(volunteer.skills,incompleteTaskObjects,requiredSkills,G_ST)

        # If no task matches the skillset of the volunteer then go for the next volunteer
        if len(suggestedTasks) == 0:
            continue
        else:
            '''
            If tasks are suggested for the given volunteer then we need to compare the most optimal task 
            based on the distance of that volunteer from each task and then assign it to the nearest task
            '''
            # selectedtask,costIncured = getCostEfficientTask(volunteer.location,list(suggestedTasks.keys()),costPerKM)
            selectedtask = getMostWillingTask(volunteer,suggestedTasks)

            if selectedtask.budget - volunteer.remuneration>=0:

                # Update the G_ST
                skillsFulfilled = suggestedTasks[selectedtask]
                updateGST(selectedtask,skillsFulfilled,requiredSkills,G_ST)


                # Update task completion info
                selectedtask.skillsLeft-=len(skillsFulfilled)
                selectedtask.budget-=volunteer.remuneration
                if selectedtask.skillsLeft == 0:
                    completedTaskObjects.add(selectedtask)
                    incompleteTaskObjects.remove(selectedtask)
                    selectedtask.finished = True


            # Update the finall output map
            volunteerTaskMap[selectedtask].append((volunteer,skillsFulfilled,volunteer.remuneration))

            assignmentsLeft[0]-=len(skillsFulfilled)
        if assignmentsLeft[0] == 0:
            return dict(volunteerTaskMap),completedTaskObjects,incompleteTaskObjects

    return dict(volunteerTaskMap),completedTaskObjects,incompleteTaskObjects


# Function to print the final VTM
def generateVTMOutput(VTMObjects):
    for vtm in VTMObjects:
        if vtm.task.finished:
            print(f"{vtm.task.name} : {[(volunteer.name,volunteer.utilisedSkills,volunteer.costIncured) for volunteer in vtm.volunteers]}")
            

# Drive VTM algorithm : O(A*T*Skills)
def driver(mainTaskObjects,taskObjects,volunteerObjects,requiredSkills,costPerKM):

    '''
    Generate G_ST                                                                                         : O(T*Skills)
    '''
    G_ST,assignmentsLeft = generateSkillTaskMapper(taskObjects,requiredSkills)


    '''
    Generate final mapping of set of volunteers to each task
    '''
    VTM,completedTaskObjects,incompleteTaskObjects = generateVolunteerTaskMap(G_ST,taskObjects,volunteerObjects,requiredSkills,costPerKM,assignmentsLeft)


    '''
    Generate the list of all VTM objects
    '''
    VolunteerTaskMapping = getVTMObjects(VTM)


    '''
    Prints the final VTM output
    '''
    generateVTMOutput(VolunteerTaskMapping)


    '''
    Generate final score for the VTM mapping.
    '''
    Scorer = ScoreVTM(VolunteerTaskMapping,len(taskObjects),volunteerObjects,mainTaskObjects)
    print(Scorer.getScores())



Tasks,NameTaskMap = getTaskObjects("Tasks_Sample.xlsx")
requiredSkills = getRequiredSkills(Tasks)
Volunteers,NameVolunteersMap = getVolunteerObjects("Applicants_Sample.xlsx",requiredSkills)
fillPreference(Tasks,NameTaskMap,Volunteers,NameVolunteersMap)


MainTaskInfo = getMainTasksObjects("MainTaskInfo_Sample.xlsx",NameTaskMap)

costPerKM = 1
driver(MainTaskInfo,Tasks,Volunteers,requiredSkills,costPerKM)