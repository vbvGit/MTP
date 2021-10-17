from openpyxl import Workbook,load_workbook
from collections import defaultdict
import pandas as pd
import time
import math
import re

'''
Future aspect : Apply dynamic behaviour to both task and applicants.
'''

'''
Data conversion functions from .xlsx to python dictionary
'''
# Convert the skills cell to list of skills and trim the extra spaces from each skill(assumes skill data in lowercase)
def preprocess_skills_xl(skill_str):
    skills = skill_str.split(',')
    fin_skill = []
    for i in range(len(skills)):
        s = re.sub(' +',' ',skills[i].strip())
        fin_skill.append(s)
    return fin_skill

# Convert the location and slot cells of excel i.e x_coord,y_coord and slot_start,slot_end
def preprocess_loc_slot_xl(x,y):
    return [int(x),int(y)]

# Convert thr Tasks.xlsx data row by row to Task dictionary 
def preprocessTaskData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    T = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row = 98,min_col = 1,max_col = 4,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        T[row[0]].append(skills)
        T[row[0]].append(loacation)
    return dict(T)

# Convert Volunteer data row by row to Applicants dictionary
def preprocessVolunteerData_xl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    A = defaultdict(list)
    for row in ws.iter_rows(min_row = 2,max_row = 1576,min_col = 1,max_col = 6,values_only = True):
        skills = preprocess_skills_xl(row[1])
        loacation = preprocess_loc_slot_xl(row[2],row[3])
        slot = preprocess_loc_slot_xl(row[4],row[5])
        A[row[0]].append(skills)
        A[row[0]].append(loacation)
        A[row[0]].append(slot)

    return dict(A)

# Preprocess the applicants data : O(AlogA)
def preprocessApplicatntSkills(A,reqSkills):
    for key,val in A.items():
        val[0][:] = [skill for skill in val[0] if reqSkills.get(skill,-1) != -1]

    A = dict(sorted(A.items(),key = lambda i : -len(i[1][0])))
    return A

def initVisistedTasks_Vw(Applicants):
    Vw = dict(zip(list(Applicants.keys()),[{} for i in range(len(Applicants))]))
    return Vw

def initResultantTeam_R_1(Tasks):
    R = {}
    for t,val in Tasks.items():
        R[t] = {}
        for skill in val[0]:
            R[t][skill] = [None,float('inf')]
    return R

def initResultantTeam_R_2(Tasks):
    R = {}
    for t,val in Tasks.items():
        R[t] = defaultdict(list)
    return R

def initAvailableApplicants(Applicants):
    availableApp = dict(zip(Applicants.keys(),[1 for i in range(len(Applicants))]))
    return availableApp

def computeCommonSkills(skillSetA,skillSetT):
    common = []
    for skill in skillSetA:
        if skill in skillSetT:
            common.append(skill)
    return common

def computeDist(locA,locB):
    return math.sqrt((locB[0]-locA[0])**2+(locB[1]-locA[1])**2)

def computeMinDistTask(locA,Tasks):
    minDistTask = None
    minDist = float('inf')
    for task,data in Tasks.items():
        dest = data[1]
        curDist = computeDist(locA,dest)
        if curDist<minDist:
            minDistTask = task
            minDist = curDist
    return minDistTask

# Drive VTM algorithm : O(A*T*Skills)
def groupMatchingAlgorithm_GMA(Tasks,Applicants,cost):

    visitedTasks_Vw = initVisistedTasks_Vw(Applicants)
    R_ResultantTeam_SkillsKey = initResultantTeam_R_1(Tasks)
    R_ResultantTeam_WorkersKey = initResultantTeam_R_2(Tasks)
    availableApplicants = initAvailableApplicants(Applicants)

    while True:
        # Gt = dict(zip(list(Tasks.keys()),[{} for i in range(len(Tasks.keys()))]))
        terminal = 1

        for applicant,availability in availableApplicants.items():
            if availableApplicants[applicant] == 1:
                minDistTask = computeMinDistTask(Applicants[applicant][1],Tasks)
                if minDistTask not in visitedTasks_Vw[applicant]:
                    # Gt[minDistTask].add(applicant)
                    visitedTasks_Vw[applicant][minDistTask] = 1

                    commonSkills = computeCommonSkills(Applicants[applicant][0],Tasks[minDistTask][0])

                    for skill in commonSkills:
                        curDist = computeDist(Tasks[minDistTask][1],Applicants[applicant][1])
                        if curDist < R_ResultantTeam_SkillsKey[minDistTask][skill][1]:
                            prevAssignedApplicant = R_ResultantTeam_SkillsKey[minDistTask][skill][0]
                            if prevAssignedApplicant is not None:
                                for s in R_ResultantTeam_WorkersKey[minDistTask][prevAssignedApplicant]:
                                    if R_ResultantTeam_SkillsKey[minDistTask][s][0] == prevAssignedApplicant:
                                        R_ResultantTeam_SkillsKey[minDistTask][s][0] = None
                                        R_ResultantTeam_SkillsKey[minDistTask][s][1] = float('inf')
                                del R_ResultantTeam_WorkersKey[minDistTask][prevAssignedApplicant]
                                availableApplicants[prevAssignedApplicant] = 1
                            R_ResultantTeam_SkillsKey[minDistTask][skill][0] = applicant
                            R_ResultantTeam_SkillsKey[minDistTask][skill][1] = curDist
                            R_ResultantTeam_WorkersKey[minDistTask][applicant].append(skill)
                            availableApplicants[applicant] = 0
                            terminal = 0

        if terminal == 1 or availableApplicants == {}:
            break

    return R_ResultantTeam_WorkersKey,R_ResultantTeam_SkillsKey




# Driver code for VTM and common slot
def driver(T,A,cost):
    # Map volunteers to tasks
    R1,R2 = groupMatchingAlgorithm_GMA(T,A,1)
    print(R1)
    print()
    print(R2)


start_time = time.time()
Tasks = preprocessTaskData_xl("Tasks.xlsx")
Applicants = preprocessVolunteerData_xl("Applicants.xlsx")
driver(Tasks,Applicants,1)
# print(f"VTM:\n{VTM}\n\nSuccess_Ratio = {success_ratio}\n\nUtility scores for all participants:\n{utilityScores}\n\nNetUtilityScore = {NetUtilityScore}\n\nTotal time taken : {time.time()-start_time}")